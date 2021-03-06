import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
import os

from produccion_gui import *


def _produccion():
    file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Producción"
    os.chdir(file_location)
    reveal = os.getcwd()
    workbook = openpyxl.load_workbook("Producción 2018.xlsx")
    sheet_1 = dia_prod
    sheet = workbook[sheet_1]

    # Procesamiento de Datos

    prod_day = int(pro_day.get())
    row = sheet.max_row + 1
    max_prod = int(max_produccion.get())
    cant_prod = int(cant_carg.get())
    max_range = row + int(max_prod)

    # ----------------------------GENERANDO ESTILOS DE CELDAS----------------------------------------
    # Formato de Celdas

    format_style = NamedStyle(name="format_style")
    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    format_style.font = Font(size=11)
    format_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    format_style.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    # workbook.add_named_style(format_style)

    # Estilo del formato inicial para la columna

    for j in range(row, max_range):
        for r in range(2, 15):
            sheet.cell(j, r).style = "format_style"
            if j == row:
                sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif j == max_range - 1:
                sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)

    # Primera Columna
            if r == 2:
                if j == row:
                    sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
                elif j == max_range - 1:
                    sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)
                else:
                    sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thin)
    # Segunda Columna
            if r == 3:
                if j == row:
                    sheet.cell(row, r).border = Border(top=thick, left=thin, right=thick, bottom=thin)
                elif j == max_range - 1:
                    sheet.cell(j, r).border = Border(top=thin, left=thin, right=thick, bottom=thick)
                else:
                    sheet.cell(j, r).border = Border(top=thin, left=thin, right=thick, bottom=thin)

    # Columna Resultados

            for r in range(4, 15):
                if j == row:
                    sheet.cell(row, r).border = Border(top=thick, left=thin, right=thin, bottom=thin)
                elif j == max_range - 1:
                    sheet.cell(j, r).border = Border(top=thin, left=thin, right=thin, bottom=thick)
                if r == 11:
                    if j == row:
                        sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
                    elif j == max_range - 1:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)
                    else:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thin)
                if r == 14:
                    if j == row:
                        sheet.cell(row, r).border = Border(top=thick, left=thick, right=thick, bottom=thin)
                    elif j == max_range - 1:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thick, bottom=thick)
                    else:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thick, bottom=thin)

    # --------------------------- DATOS CASILLAS EXCEL ----------------------------------------------
    # Lote de Procesamiento
        if prod_day < 10:
            sheet.cell(j, 2).value = 'L' + '0' + str(prod_day) + str(sheet_prov) + '18'
        else:
            sheet.cell(j, 2).value = 'L' + str(prod_day) + str(sheet_prov) + '18'
        sheet.cell(j, 3).value = '0' + str(j - (row - 1))

    # Carga alimentada + cantidad no alimentada

        sheet.cell(j, 4).value = int(cant_prod)
        sheet.cell(j, 5).value = 0.03

    # Valores Fracción Gruesa
    # VALOR A CAMBIAR FRACCION GRUESA input("Ingrese producción fracción gruesa {}: "
        sheet.cell(j, 6).value = int((entries[j - (row - 1)].get()))

    # Valor Resultados

    for co in range(7, 15):
        if co != 8:
            if co == 7:
                sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                sheet.cell(row, co).value = int(cant_fina.get())
            if co == 9:
                sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                sheet.cell(row, co).value = '=SUM($F${0}:$G${1})'.format(row, (max_range - 1))
            if co == 10:
                sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                sheet.cell(row, co).value = '=($D${0}*$C${1})-$I${0}'.format(row, (max_range - 1))
            if co == 11:
                sheet.merge_cells(start_row=row, start_column=co, end_row=(max_range - 1), end_column=13)
                sheet.cell(row, co).value = str(obs.get())
            if co == 14:
                sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                sheet.cell(row, co).value = '=SUM($F${0}:$F${1})'.format(row, (max_range - 1))
        else:
            continue
    boton_act = tk.Button(dialogo2, text='Actualizar BD', command=_produccion())
    boton_act.grid(row=3, column=5, padx=10)
    workbook.save('Producción 2018.xlsx')
