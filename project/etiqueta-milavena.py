import openpyxl
import os
from openpyxl.styles import PatternFill


def milavena(self):

    file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Productos/Documentación productos CCF/Etiquetado"
    os.chdir(file_location)
    reveal = os.getcwd()
    workbook = openpyxl.load_workbook('Base de datos_Milavena.xlsx')
    sheet = workbook['BD']
    sheet.insert_rows(2, 1)
    column = sheet.max_column
    row = sheet.max_row
    print('Procesando...')
    for j in range(1, column + 1):
        sheet.cell(2, j).value = sheet.cell(3, j).value
        if j == 7:
            sheet.cell(2, j).value = 'CCF0000' + input('Ingrese número de Lote: ')
        if j == 8:
            date_elab = input("Ingrese fecha de elaboración (dd/mm/aa): ")
            d = date_elab[2:4]
            print(d)
            month = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
            month_date = month[int(float(d)) - 1]
            sheet.cell(2, j).value = date_elab[0:2] + '-' + month_date + '-' + '20' + date_elab[4:6]
            yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            sheet.cell(2, j).fill = yellowFill
        if j == 14:
            sheet.cell(2, j).value = int(input('Ingrese cantidad de etiquetas a imprimir: '))
        sheet.cell(2, 15).value = sheet.cell(2, 14).value

    workbook.save('Base de datos_Milavena.xlsx')
    print('Etiquetas ingresadas!')
