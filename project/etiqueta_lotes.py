import openpyxl
import os
import tkinter as tk
from tkinter import ttk
import subprocess


class LOT():

    def _lotes(self):

        self.dialogo2 = tk.Toplevel()
        self.dialogo2.geometry('570x230+350+150')
        self.dialogo2.title('Etiqueta Fibravena')
        customer_frame = tk.LabelFrame(self.dialogo2, text='Etiqueta Lotes', font='Sunday', height=2, fg='#FFFFFF', bg='#5475B1', bd=0)
        customer_frame.grid(row=0, column=0, padx=10, pady=20)

        # Número de Lote
        self.max_prod = tk.IntVar()

        def onClick():
            self.max_prod.set(self.max_prod.get() + 1)
            self.cant_etql = []
            self.cant_tipo = []
            self.cant_lotter = []
            for i in range(self.max_prod.get()):
                self.label_prod = ttk.Combobox(customer_frame, values=['Milavena', 'Protavena', 'Fibravena'], state='readonly')
                self.label_prod.grid(row=i + 2, column=0)
                self.cant_tipo.append(self.label_prod)
                self.c = tk.Entry(customer_frame, width=10)
                self.c.grid(row=i + 2, column=1)
                self.cant_etql.append(self.c)

                def max_lotes(event):
                    self.cant_lotter.append(self.cant_etq.get())
                    print(self.cant_lotter)
                self.cant_etq = ttk.Combobox(customer_frame, values=[x for x in range(1, 17)], state='readonly')
                self.cant_etq.grid(row=i + 2, column=2)
                self.cant_lotter.append(self.cant_etq)

        def action(event):
            for i in range(self.max_prod.get()):
                self.label_act = tk.Label(customer_frame, text='✔', fg='#FFFFFF', bg='#5475B1')
                self.label_act.grid(row=i + 1, column=3)

        def setText():
            self.entries = []
            self.entries2 = []
            self.entries3 = []
            print(self.cant_lotter)
            for self.c in self.cant_etql:
                self.entries.append(self.c.get())
            for self.label_prod in self.cant_tipo:
                self.entries2.append(self.label_prod.get())
            for self.cant_etq in self.cant_lotter:
                self.entries3.append(self.cant_etq.get())

        etiq = tk.Label(customer_frame, text='Etiquetas a Imprimir', fg='#FFFFFF', bg='#151763', width=25).grid(row=0, column=0, padx=10, pady=10)
        label_tlot = tk.Label(customer_frame, text='Tipo de Producto', fg='#FFFFFF', bg='#151763', width=20)
        label_tlot.grid(row=1, column=0)
        label_lot = tk.Label(customer_frame, text='Lote', fg='#FFFFFF', bg='#151763', width=10)
        label_lot.grid(row=1, column=1)
        label_cant = tk.Label(customer_frame, text='Cantidad', fg='#FFFFFF', bg='#151763', width=10)
        label_cant.grid(row=1, column=2)
        lotes = tk.Button(customer_frame, text='+', command=onClick, fg='gray24', bg='papaya whip')
        lotes.grid(row=0, column=1, stick='e')
        boton_max = tk.Button(customer_frame, text='Set', command=setText, fg='gray24', bg='papaya whip')
        boton_max.grid(row=0, column=2, pady=10, padx=10)
        boton_max.bind('<Button-1>', action)

        # Actualizar BD
        customer_frame3 = tk.LabelFrame(self.dialogo2, bg='#FFFFFF', bd=2)
        customer_frame3.grid(row=0, column=2, padx=10, pady=20)

        botonBD = tk.Button(customer_frame3, text='Actualizar BD', command=self.base_etq, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=10)
        botonBD.grid(row=0, column=2, pady=10, padx=10)
        # Imprimir
        botonIMP = tk.Button(customer_frame3, text='Imprimir', command=self.goLabel, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=10)
        botonIMP.grid(row=1, column=2, pady=10, padx=10)
        # Cerrar
        botonc = tk.Button(customer_frame3, text='Cerrar', command=self.dialogo2.destroy, bd=0, fg='#FFFFFF', bg='#151763', height=2, width=10)
        botonc.grid(row=2, column=2, padx=10, pady=10)
        self.dialogo2.grab_set()
        self.dialogo2.wait_window(self.dialogo2)

    def base_etq(self):

        file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Base de Datos/Etiquetado"
        os.chdir(file_location)
        workbook = openpyxl.load_workbook('Base de datos_Lotes.xlsx')
        sheet = workbook['BD']
        sheet.insert_rows(2, 1)
        intervalo = 1
        print(self.entries3)
        print(self.entries2)
        print(self.entries)
        for r in range(1, self.max_prod.get() + 1):
            maxintervalo = int(self.entries3[r - 1])
            for j in range(intervalo, (maxintervalo + intervalo)):
                if self.entries2[r - 1] == 'Milavena':
                    sheet.cell(2, j).value = 'CCF0000' + str(self.entries[r - 1])
                if self.entries2[r - 1] == 'Protavena':
                    sheet.cell(2, j).value = 'CCF0000' + str(self.entries[r - 1])
                if self.entries2[r - 1] == 'Fibravena':
                    sheet.cell(2, j).value = 'CCF13000' + str(self.entries[r - 1])
            intervalo = int(self.entries3[r - 1]) + 1
        workbook.save('Base de datos_Lotes.xlsx')

    def goLabel(self):
        cmd = "C:/Program Files (x86)/GoDEX/GoLabel/GoLabel.exe"
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, creationflags=0x08000000)
        process.wait()
