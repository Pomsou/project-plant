import openpyxl
import os
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk
import datetime as dt
import subprocess


class PROT():

    def protavena(self):

        self.dialogo2 = tk.Toplevel()
        self.dialogo2.geometry('520x230+350+150')
        self.dialogo2.title('Etiqueta Protavena')
        customer_frame = tk.LabelFrame(self.dialogo2, text='Etiqueta Protavena', font='Sunday', height=2, fg='#FFFFFF', bg='#5475B1', bd=0)
        customer_frame.grid(row=0, column=0, padx=10, pady=20)
        self.label_lotes = tk.Label(customer_frame, text='Número de Lote: ', fg='#FFFFFF', bg='#151763', width=20)
        self.label_lotes.grid(row=1, column=0, padx=10, pady=10)
        self.lotes = tk.Entry(customer_frame, width=7)
        self.lotes.grid(row=1, column=1, pady=10, stick='w')
        # dia
        fecha = tk.Label(customer_frame, text='Fecha de Elaboración', fg='#FFFFFF', bg='#151763', width=20)
        fecha.grid(row=2, column=0, padx=10, pady=10)

        def check_day(event):
            self.pro_day = day.get()
        day = ttk.Combobox(customer_frame, values=[i for i in range(1, 32)], state='readonly')
        day.grid(row=3, column=0, padx=10, pady=2)
        day.bind('<<ComboboxSelected>>', check_day)

        # mes
        def check_prod(event):
            self.mes_value = str(mes.get())

        mes = ttk.Combobox(customer_frame, values=('enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'), state='readonly')
        mes.grid(row=3, column=1, padx=10, pady=2)
        mes.bind("<<ComboboxSelected>>", check_prod)

        # Año
        today = dt.date.today()
        self.year = today.year

        # Cantidad de Etiquetas
        cant_etq = tk.Label(customer_frame, text='Cantidad a imprimir:', fg='#FFFFFF', bg='#151763', width=20)
        cant_etq.grid(row=4, column=0, padx=10, pady=10)
        self.cantetq = tk.Entry(customer_frame, width=7)
        self.cantetq.grid(row=4, column=1, pady=10, stick='w')

        # Actualizar BD
        customer_frame2 = tk.LabelFrame(self.dialogo2, bg='#FFFFFF', bd=2)
        customer_frame2.grid(row=0, column=2, padx=10, pady=20)

        botonBD = tk.Button(customer_frame2, text='Actualizar BD', command=self.base_etq, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=10)
        botonBD.grid(row=0, column=2, pady=10, padx=10)
        # Imprimir
        botonIMP = tk.Button(customer_frame2, text='Imprimir', command=self.goLabel, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=10)
        botonIMP.grid(row=1, column=2, pady=10, padx=10)
        # Cerrar
        botonc = tk.Button(customer_frame2, text='Cerrar', command=self.dialogo2.destroy, bd=0, fg='#FFFFFF', bg='#151763', height=2, width=10)
        botonc.grid(row=2, column=2, padx=10, pady=10)
        self.dialogo2.grab_set()
        self.dialogo2.wait_window(self.dialogo2)

    def base_etq(self):

        file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Productos/Documentación productos CCF/Etiquetado"
        os.chdir(file_location)
        workbook = openpyxl.load_workbook('Base de datos_Protavena.xlsx')
        sheet = workbook['BD']
        sheet.insert_rows(2, 1)
        column = sheet.max_column
        row = sheet.max_row
        cant_lotes = int(self.lotes.get())
        dia_prod = int(self.pro_day)
        cant_etiquetas = int(self.cantetq.get())
        for j in range(1, column + 1):
            sheet.cell(2, j).value = sheet.cell(3, j).value
            if j == 7:
                sheet.cell(2, j).value = 'CCF0000' + str(cant_lotes)
            if j == 8:
                month_date = self.mes_value[0:3]
                if dia_prod < 10:
                    sheet.cell(2, j).value = '0' + str(dia_prod) + '-' + month_date + '-' + str(self.year)
                else:
                    sheet.cell(2, j).value = str(dia_prod) + '-' + month_date + '-' + str(self.year)
                yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                sheet.cell(2, j).fill = yellowFill
            if j == 14:
                sheet.cell(2, j).value = cant_etiquetas
            sheet.cell(2, 15).value = sheet.cell(2, 14).value
        # Reconocimiento dia de elaboración anterior
        for i in range(1, row + 1):
            if 'CCF0000{0}'.format(str(cant_lotes)) == sheet.cell(i, 7).value:
                sheet.cell(2, 8).value = sheet.cell(i, 8).value
        workbook.save('Base de datos_Protavena.xlsx')

    def goLabel(self):
        cmd = "C:/Program Files (x86)/GoDEX/GoLabel/GoLabel.exe"
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, creationflags=0x08000000)
        process.wait()
