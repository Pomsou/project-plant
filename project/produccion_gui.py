import tkinter as tk
from tkinter import Button, ttk
import tkinter.scrolledtext as tsk
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
import os
import datetime
import subprocess
import psutil


class PRD:
    file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Producción"
    os.chdir(file_location)

    def produccion(self):
        self.dialogo = tk.Toplevel()
        self.dialogo.geometry('570x400+350+140')
        self.dialogo.resizable(0, 0)
        self.dialogo.title('Producción')
        # Design UI
        self.TopFrame = tk.Frame(self.dialogo, bg='#151763', bd=0, width=570, height=100)
        self.TopFrame.grid(column=0, row=0, ipadx=10, ipady=10, stick='we')

        # Registro Producción Mensual
        customer_prom = tk.LabelFrame(self.dialogo, bg='#151763', bd=0)
        customer_prom.grid(row=1, column=0, pady=10)
        etiq_prom = tk.Label(customer_prom, text='Producción Mes Actual', font='Sunday', height=3, fg='#FFFFFF', bg='#151763')
        etiq_prom.grid(row=1, column=0, stick='e', padx=10, pady=10)
        self.file_mes = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        self.fecha_hoy = datetime.date.today()
        self.año = self.fecha_hoy.year
        self.mes2 = self.fecha_hoy.month
        wb = openpyxl.load_workbook("Producción {0}.xlsx".format(self.año), data_only=True)
        sheet = wb[self.file_mes[self.mes2 - 2]]
        reg_milavena = sheet.cell(5, 17).value
        reg_protavena = sheet.cell(6, 17).value
        result_merma = sheet.cell(7, 17).value
        label_mil = tk.Label(customer_prom, text='Milavena {0:.2f} kg'.format(float(reg_milavena)), width=15, height=2)
        label_mil.grid(row=2, column=0, stick='we', padx=10)
        label_prot = tk.Label(customer_prom, text='Protavena {0:.2f} kg'.format(float(reg_protavena)), width=15, height=2)
        label_prot.grid(row=3, column=0, stick='we', padx=10, pady=10)

        # Registro Producción Mes Anterior
        etiq_prom2 = tk.Label(customer_prom, text='Producción Mes Anterior', font='Sunday', height=3, fg='#FFFFFF', bg='#151763')
        etiq_prom2.grid(row=1, column=2, stick='e', padx=10, pady=10)
        sheet2 = wb[self.file_mes[self.mes2 - 2]]
        reg_milavena2 = sheet2.cell(5, 17).value
        reg_protavena2 = sheet2.cell(6, 17).value
        label_mil2 = tk.Label(customer_prom, text='Milavena {0:.2f} kg'.format(float(reg_milavena2)), width=15, height=2)
        label_mil2.grid(row=2, column=2, stick='we', padx=10)
        label_prot2 = tk.Label(customer_prom, text='Protavena {0:.2f} kg'.format(float(reg_protavena2)), width=15, height=2)
        label_prot2.grid(row=3, column=2, stick='we', padx=10)

        # Botón Cerrar y mantener window inicial
        boton = Button(self.dialogo, text='Cerrar', command=self.dialogo.destroy, bd=0, fg='#FFFFFF', bg='#151763', height=2, width=10)
        boton.grid(row=5, column=0, pady=10)
        botonp = Button(self.dialogo, text='Agregar producción diaria', command=PRD().prod_diaria, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=20)
        botonp.grid(row=4, column=0)
        self.dialogo.transient()
        self.dialogo.grab_set()
        self.dialogo.wait_window(self.dialogo)

    def prod_diaria(self):
        self.dialogo2 = tk.Toplevel()
        self.dialogo2.geometry('700x400+340+120')
        self.dialogo2.resizable(0, 0)
        self.dialogo2.title('Producción diaria')
        # Label producción
        customer_frame = tk.LabelFrame(self.dialogo2, text='Fecha', font='Sunday', height=2, fg='#FFFFFF', bg='#5475B1', bd=0)
        customer_frame.grid(row=0, column=0, padx=10, pady=20)
        customer_frame2 = tk.LabelFrame(self.dialogo2, text='Producción', font='Sunday', height=2, fg='#FFFFFF', bg='#5475B1', bd=0)
        customer_frame2.grid(row=0, column=3, padx=10, pady=10)
        customer_frame3 = tk.LabelFrame(self.dialogo2, text='Observación', font='Sunday', height=2, fg='#FFFFFF', bg='#5475B1', bd=0)
        customer_frame3.grid(row=1, column=0, rowspan=4, columnspan=5, sticky='W' + 'E' + 'N' + 'S', padx=5, pady=5)

        # dia de producción
        def check_day(event):
            self.pro_day = day.get()
        etiq2 = tk.Label(customer_frame, text='Dia de Producción:', fg='#FFFFFF', bg='#151763')
        etiq2.grid(row=0, column=0, stick='we', padx=10, pady=10)
        day = ttk.Combobox(customer_frame, values=[i for i in range(1, 32)], state='readonly')
        day.grid(row=1, column=0, padx=10, pady=2)
        day.bind('<<ComboboxSelected>>', check_day)

        # Mes de producción
        def check_prod(event):
            self.values = ('Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre')
            self.dia_prod = mes.get()
            self.mes_value = self.values.index(self.dia_prod) + 1
        etiq1 = tk.Label(customer_frame, text='Mes de Producción:', fg='#FFFFFF', bg='#151763')
        etiq1.grid(row=2, column=0, stick='we', padx=10, pady=10)
        mes = ttk.Combobox(customer_frame, values=('Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'), state='readonly')
        mes.grid(row=3, column=0, padx=10, pady=2)
        mes.bind("<<ComboboxSelected>>", check_prod)
        # Año de producción
        year_label = tk.Label(customer_frame, text='Año', fg='#FFFFFF', bg='#151763', width=13)
        year_label.grid(row=4, column=0, pady=10, padx=10, stick='w')
        self.year_date = tk.Entry(customer_frame, width=5)
        self.year_date.grid(row=4, column=0, stick='e', pady=10, padx=5)

        # Carga Alimentada
        etiqc = tk.Label(customer_frame, text='Carga Alimentada', fg='#FFFFFF', bg='#151763').grid(row=5, column=0, stick='w', padx=10, pady=5)
        self.cantidad_prod = tk.Entry(customer_frame, width=5)
        self.cantidad_prod.grid(row=5, column=0, padx=5, pady=10, stick='e')

        # Número de lotes
        self.max_prod = tk.IntVar()

        def onClick():
            self.max_prod.set(self.max_prod.get() + 1)
            self.fr_gruesa = []
            for i in range(self.max_prod.get()):
                label_prod = tk.Label(customer_frame2, text='%2d.-' % (i + 1), fg='#FFFFFF', bg='#5475B1')
                label_prod.grid(row=i + 1, column=2)
                self.c = tk.Entry(customer_frame2, width=10)
                self.c.grid(row=i + 1, column=3)
                self.fr_gruesa.append(self.c)

        def action(event):
            for i in range(self.max_prod.get()):
                self.label_act = tk.Label(customer_frame2, text='✔', fg='#FFFFFF', bg='#5475B1')
                self.label_act.grid(row=i + 1, column=4)

        def action2(event):
            label_tik = tk.Label(customer_frame2, text='✔', fg='#FFFFFF', bg='#5475B1')
            label_tik.grid(row=11, column=4)

        def action3(event):
            label_obs = tk.Label(customer_frame3, text='✔', fg='#FFFFFF', bg='#5475B1')
            label_obs.grid(row=3, column=5)

        def action4(event):
            label_cg = tk.Label(customer_frame, text='✔', fg='#FFFFFF', bg='#5475B1')
            label_cg.grid(row=5, column=1)

        # Cantidad de produccion + label

        def setText():
            self.entries = []
            print(self.fr_gruesa)
            for self.c in self.fr_gruesa:
                self.entries.append(self.c.get())

        etiq4 = tk.Label(customer_frame2, text='Producción Fracción Gruesa', fg='#FFFFFF', bg='#151763', width=25).grid(row=0, column=3, padx=10, pady=10)
        lotes = tk.Button(customer_frame2, text='+', command=onClick, fg='gray24', bg='papaya whip')
        lotes.grid(row=0, column=4)
        boton_max = tk.Button(customer_frame2, text='Set', command=setText, fg='gray24', bg='papaya whip')
        boton_max.grid(row=0, column=5, pady=10)
        boton_max.bind('<Button-1>', action)
        # Cantidad de producción fina
        etiq5 = tk.Label(customer_frame2, text='Producción Fracción Fina', fg='#FFFFFF', bg='#151763', width=25).grid(row=10, column=3, padx=10, pady=5)
        self.cant_fina = tk.Entry(customer_frame2, width=10)
        self.cant_fina.grid(row=11, column=3, pady=10)

        # Observaciones
        etiq6 = tk.Label(customer_frame3, text='Observaciones').grid(row=3, column=0, rowspan=4, columnspan=5, sticky='W' + 'E' + 'N' + 'S', padx=5, pady=5)
        self.obs = tsk.ScrolledText(master=customer_frame3, wrap=tk.WORD, width=60, height=4)
        self.obs.grid(row=3, column=0, rowspan=4, columnspan=5, sticky='W' + 'E' + 'N' + 'S', padx=10, pady=10)
        self.obs.bind('<Return><Return>', action3)

        # Actualizar Base de datos
        boton_act = tk.Button(self.dialogo2, text='Actualizar BD', command=self._produccion, bd=0, fg='#FFFFFF', bg='#5475B1', height=2, width=20)
        boton_act.grid(row=1, column=5, padx=10)

        # Apertura archivo
        def abrir_file():
            fecha_hoy = datetime.date.today()
            año = fecha_hoy.year
            doc = "Producción {0}.xlsx".format(año)
            os.startfile(doc)

        abrir_boton = Button(self.dialogo2, text='Abrir archivo', command=abrir_file, bd=0, fg='#FFFFFF', bg='#151763', height=2, width=20)
        abrir_boton.grid(row=2, column=5, padx=10, pady=5)
        # Apertura y guardado de archivo
        self.file_location = "//Gcl-file-sr/CCF/03 Planta de Fraccionamiento/Operativo/Producción"
        os.chdir(self.file_location)
        reveal = os.getcwd()

        def check_file():
            self.year = self.year_date.get()
            if os.path.isfile('Producción {0}.xlsx'.format(self.year)):
                self.workbook = openpyxl.load_workbook("Producción {0}.xlsx".format(self.year))
            else:
                self.workbook = openpyxl.load_workbook('Tipo.xlsx')
                self.format_style = NamedStyle(name="format_style")
                thick = Side(border_style="medium", color="000000")
                thin = Side(border_style="thin", color="000000")
                self.format_style.font = Font(size=11)
                self.format_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                self.format_style.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
                self.workbook.add_named_style(self.format_style)
                self.workbook.save('Producción {0}.xlsx'.format(self.year))

        file_boton = tk.Button(customer_frame, text='Revisión', command=check_file)
        file_boton.grid(row=4, column=1, stick='e', padx=5, pady=10)

        # Botón Cerrar y mantener window inicial
        boton = Button(self.dialogo2, text='Cerrar', command=self.dialogo2.destroy, bd=0, fg='#FFFFFF', bg='#151763', height=2, width=20)
        boton.grid(row=3, column=5, padx=10)
        self.dialogo2.grab_set()
        self.dialogo2.wait_window(self.dialogo2)

    def _produccion(self):
        sheet_1 = self.dia_prod
        sheet_2 = self.mes_value
        # Creación de pestaña por mes (si no existe)
        if sheet_1 in self.workbook.sheetnames:
            sheet = self.workbook[sheet_1]

        else:
            sheet = self.workbook.copy_worksheet(self.workbook['Hoja1'])
            sheet.title = sheet_1

            self.workbook.save('Producción {0}.xlsx'.format(self.year))

        # Procesamiento de Datos
        row = sheet.max_row + 1
        prod_day = int(self.pro_day)
        cant_prod = int(self.cantidad_prod.get())
        max_range = row + int(self.max_prod.get())

        # Registro Producción Fino
        sheet.cell(5, 17).value = '=SUM($G${0}:$G${1})'.format(5, row + 1)

        # Registro Producción Grueso
        sheet.cell(6, 17).value = '=SUM($N${0}:$N${1})'.format(5, row + 1)

        # Registro Merma
        sheet.cell(7, 17).value = '=SUM($J${0}:$J${1})'.format(5, row + 1)
        # Formato de Celdas

        thick = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # xlwing automatizar fórmulas

        # Estilo del formato inicial para la columna

        for j in range(row, max_range):
            for r in range(2, 15):
                sheet.cell(j, r).style = 'format_style'
                if j == row:
                    sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
                elif j == max_range - 1:
                    sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)

        # Primera Columna
                if r == 2:
                    if j == row:
                        sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
                    elif j == max_range - 1:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)
                    else:
                        sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thin)
        # Segunda Columna
                if r == 3:
                    if j == row:
                        sheet.cell(row, r).border = Border(top=thick, left=thin, right=thick, bottom=thin)
                    elif j == max_range - 1:
                        sheet.cell(j, r).border = Border(top=thin, left=thin, right=thick, bottom=thick)
                    else:
                        sheet.cell(j, r).border = Border(top=thin, left=thin, right=thick, bottom=thin)

        # Columna Resultados

                for r in range(4, 15):
                    if j == row:
                        sheet.cell(row, r).border = Border(top=thick, left=thin, right=thin, bottom=thin)
                    elif j == max_range - 1:
                        sheet.cell(j, r).border = Border(top=thin, left=thin, right=thin, bottom=thick)
                    if r == 11:
                        if j == row:
                            sheet.cell(row, r).border = Border(top=thick, left=thick, right=thin, bottom=thin)
                        elif j == max_range - 1:
                            sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thick)
                        else:
                            sheet.cell(j, r).border = Border(top=thin, left=thick, right=thin, bottom=thin)
                    if r == 14:
                        if j == row:
                            sheet.cell(row, r).border = Border(top=thick, left=thick, right=thick, bottom=thin)
                        elif j == max_range - 1:
                            sheet.cell(j, r).border = Border(top=thin, left=thick, right=thick, bottom=thick)
                        else:
                            sheet.cell(j, r).border = Border(top=thin, left=thick, right=thick, bottom=thin)

        # --------------------------- DATOS CASILLAS EXCEL ----------------------------------------------
        # Lote de Procesamiento
            if (prod_day < 10 and sheet_2 < 10):
                sheet.cell(j, 2).value = 'L' + '0' + str(prod_day) + '0' + str(sheet_2) + '18'
            elif prod_day < 10:
                sheet.cell(j, 2).value = 'L' + '0' + str(prod_day) + str(sheet_2) + '18'
            elif sheet_2 < 10:
                sheet.cell(j, 2).value = 'L' + str(prod_day) + '0' + str(sheet_2) + '18'
            else:
                sheet.cell(j, 2).value = 'L' + str(prod_day) + str(sheet_2) + '18'
            sheet.cell(j, 3).value = '0' + str(j - (row - 1))

        # Carga alimentada + cantidad no alimentada

            sheet.cell(j, 4).value = int(cant_prod)
            sheet.cell(j, 5).value = 0.03

        # Valores Fracción Gruesa
        # VALOR A CAMBIAR FRACCION GRUESA input("Ingrese producción fracción gruesa {}: "
            sheet.cell(j, 6).value = int(str(self.entries[(j - row)]))

        # Valor Resultados

        for co in range(7, 15):
            if co != 8:
                if co == 7:
                    sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                    sheet.cell(row, co).value = int(self.cant_fina.get())
                if co == 9:
                    sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                    sheet.cell(row, co).value = '=SUM($F${0}:$G${1})'.format(row, (max_range - 1))
                if co == 10:
                    sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                    sheet.cell(row, co).value = '=($D${0}*$C${1})-$I${0}'.format(row, (max_range - 1))
                if co == 11:
                    sheet.merge_cells(start_row=row, start_column=co, end_row=(max_range - 1), end_column=13)
                    sheet.cell(row, co).value = str(self.obs.get('1.0', 'end-2c'))
                if co == 14:
                    sheet.merge_cells(start_column=co, start_row=row, end_row=(max_range - 1), end_column=co)
                    sheet.cell(row, co).value = '=SUM($F${0}:$F${1})'.format(row, (max_range - 1))
            else:
                continue

        self.workbook.save('Producción {0}.xlsx'.format(self.year))
